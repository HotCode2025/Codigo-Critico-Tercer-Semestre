"""
Código Crítico - Tercer Semestre Año 2026
Generador de Reportes.
Permite exportar los reportes predefinidos a PDF y Excel,
tomando la configuración de logo/encabezados desde la tabla 'parametros'.
"""

import sqlite3
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Table, TableStyle,
                                Image, Spacer, PageBreak)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from controladores.controlador_reportes import ControladorReportes

class GeneradorReportes:
    def __init__(self, db: sqlite3.Connection):
        self.db = db
        self.ctrl_reportes = ControladorReportes(db)
        # Cargar parámetros de la empresa
        cur = self.db.cursor()
        cur.execute("SELECT * FROM parametros WHERE id = 1")
        row = cur.fetchone()
        self.params = dict(row) if row else {}
        self.logo_path = self._guardar_logo_temporal()

    def _guardar_logo_temporal(self) -> str:
        """Extrae el logo de la base de datos y lo guarda en un archivo temporal."""
        logo_blob = self.params.get('logo')
        if logo_blob:
            temp_path = os.path.join(os.path.dirname(__file__), "..", "assets", "logo_temp.png")
            os.makedirs(os.path.dirname(temp_path), exist_ok=True)
            with open(temp_path, 'wb') as f:
                f.write(logo_blob)
            return temp_path
        return ""

    def _estilo_encabezado(self):
        """Devuelve estilos comunes para los reportes PDF."""
        styles = getSampleStyleSheet()
        estilo_titulo = ParagraphStyle(
            'TituloReporte',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=12
        )
        estilo_empresa = ParagraphStyle(
            'DatosEmpresa',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER
        )
        return styles, estilo_titulo, estilo_empresa

    def _agregar_encabezado_pdf(self, elementos, titulo_reporte: str):
        """Inserta el logo, nombre de empresa y título del reporte."""
        estilos, estilo_titulo, estilo_empresa = self._estilo_encabezado()

        if self.logo_path and os.path.exists(self.logo_path):
            try:
                logo = Image(self.logo_path, width=60*mm, height=20*mm)
                logo.hAlign = 'CENTER'
                elementos.append(logo)
            except:
                pass

        nombre = self.params.get('nombre_distribuidora', 'Distribuidora')
        tel = self.params.get('telefono1', '')
        email = self.params.get('email', '')
        encabezado_fact = self.params.get('encabezado_factura', '')
        elementos.append(Paragraph(nombre, estilo_empresa))
        if tel or email:
            elementos.append(Paragraph(f"Tel: {tel}  |  Email: {email}", estilo_empresa))
        if encabezado_fact:
            elementos.append(Paragraph(encabezado_fact, estilo_empresa))

        elementos.append(Spacer(1, 10*mm))
        elementos.append(Paragraph(titulo_reporte, estilo_titulo))
        elementos.append(Spacer(1, 5*mm))

    def _data_a_tabla(self, datos: list, columnas: list) -> list:
        """Convierte una lista de diccionarios en una lista de listas para la tabla."""
        tabla = [columnas]
        for fila in datos:
            tabla.append([str(fila.get(col, '')) for col in columnas])
        return tabla

    def _estilo_tabla_basico(self, tabla_datos):
        """Aplica estilo simple a una tabla."""
        estilo = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ])
        return estilo

    # ------------------- Reportes PDF -------------------
    def generar_pdf_productos_mas_vendidos(self, ruta_salida: str):
        """Genera PDF con productos más vendidos por mes."""
        doc = SimpleDocTemplate(ruta_salida, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=20*mm, bottomMargin=20*mm)
        elementos = []
        self._agregar_encabezado_pdf(elementos, "Productos más vendidos por mes")

        datos = self.ctrl_reportes.productos_mas_vendidos_por_mes()
        columnas = ["Producto", "Mes", "Cantidad Vendida"]
        tabla = self._data_a_tabla(datos, columnas)

        t = Table(tabla, repeatRows=1)
        t.setStyle(self._estilo_tabla_basico(tabla))
        elementos.append(t)

        doc.build(elementos)
        return ruta_salida

    def generar_pdf_cc_al_limite(self, ruta_salida: str):
        """PDF con clientes cuyo saldo supera el 80% del límite."""
        doc = SimpleDocTemplate(ruta_salida, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=20*mm, bottomMargin=20*mm)
        elementos = []
        self._agregar_encabezado_pdf(elementos, "Clientes con Cuenta Corriente al Límite")

        datos = self.ctrl_reportes.clientes_con_cc_al_limite()
        columnas = ["Razón Social", "CUIT", "Límite", "Saldo CC", "% Utilizado"]
        tabla = self._data_a_tabla(datos, columnas)

        t = Table(tabla, repeatRows=1)
        t.setStyle(self._estilo_tabla_basico(tabla))
        elementos.append(t)

        doc.build(elementos)
        return ruta_salida

    def generar_pdf_ganancia_productos(self, ruta_salida: str):
        """PDF con ganancias por producto (costo vs venta)."""
        doc = SimpleDocTemplate(ruta_salida, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=20*mm, bottomMargin=20*mm)
        elementos = []
        self._agregar_encabezado_pdf(elementos, "Reporte de Ganancias por Producto")

        datos = self.ctrl_reportes.ganancia_por_producto()
        columnas = ["Producto", "Costo", "Venta Prom.", "Cantidad", "Ganancia Total"]
        tabla = self._data_a_tabla(datos, columnas)

        t = Table(tabla, repeatRows=1)
        t.setStyle(self._estilo_tabla_basico(tabla))
        elementos.append(t)

        doc.build(elementos)
        return ruta_salida

    def generar_pdf_deuda_sin_cobrar(self, ruta_salida: str):
        """PDF con mercadería vendida y no cobrada."""
        doc = SimpleDocTemplate(ruta_salida, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=20*mm, bottomMargin=20*mm)
        elementos = []
        self._agregar_encabezado_pdf(elementos, "Mercadería Vendida sin Cobrar")

        datos = self.ctrl_reportes.mercaderia_vendida_sin_cobrar()
        columnas = ["Cliente", "Factura", "Total", "Pendiente"]
        tabla = self._data_a_tabla(datos, columnas)

        t = Table(tabla, repeatRows=1)
        t.setStyle(self._estilo_tabla_basico(tabla))
        elementos.append(t)

        doc.build(elementos)
        return ruta_salida

    def generar_pdf_ventas_por_preventista(self, ruta_salida: str):
        """PDF con ventas totales por preventista."""
        doc = SimpleDocTemplate(ruta_salida, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=20*mm, bottomMargin=20*mm)
        elementos = []
        self._agregar_encabezado_pdf(elementos, "Ventas por Preventista")

        datos = self.ctrl_reportes.ventas_por_preventista()
        columnas = ["Preventista", "Cant. Facturas", "Total Vendido"]
        tabla = self._data_a_tabla(datos, columnas)

        t = Table(tabla, repeatRows=1)
        t.setStyle(self._estilo_tabla_basico(tabla))
        elementos.append(t)

        doc.build(elementos)
        return ruta_salida

    # ------------------- Reportes Excel -------------------
    def _exportar_excel(self, datos: list, columnas: list, ruta_salida: str, titulo: str):
        """Método genérico para exportar a Excel con openpyxl."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("Se requiere openpyxl para exportar a Excel. Instale con: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = titulo[:31]  # máximo 31 caracteres
        # Encabezados
        ws.append(columnas)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        # Datos
        for fila in datos:
            ws.append([fila.get(col, '') for col in columnas])
        wb.save(ruta_salida)
        return ruta_salida

    def generar_excel_productos_mas_vendidos(self, ruta_salida: str):
        datos = self.ctrl_reportes.productos_mas_vendidos_por_mes()
        columnas = ["Producto", "Mes", "Cantidad Vendida"]
        return self._exportar_excel(datos, columnas, ruta_salida, "Prod. x Mes")

    def generar_excel_cc_al_limite(self, ruta_salida: str):
        datos = self.ctrl_reportes.clientes_con_cc_al_limite()
        columnas = ["Razón Social", "CUIT", "Límite", "Saldo CC", "% Utilizado"]
        return self._exportar_excel(datos, columnas, ruta_salida, "CC al Límite")

    def generar_excel_ganancia_productos(self, ruta_salida: str):
        datos = self.ctrl_reportes.ganancia_por_producto()
        columnas = ["Producto", "Costo", "Venta Prom.", "Cantidad", "Ganancia Total"]
        return self._exportar_excel(datos, columnas, ruta_salida, "Ganancia Prod.")

    def generar_excel_deuda_sin_cobrar(self, ruta_salida: str):
        datos = self.ctrl_reportes.mercaderia_vendida_sin_cobrar()
        columnas = ["Cliente", "Factura", "Total", "Pendiente"]
        return self._exportar_excel(datos, columnas, ruta_salida, "Deuda sin Cobrar")

    def generar_excel_ventas_por_preventista(self, ruta_salida: str):
        datos = self.ctrl_reportes.ventas_por_preventista()
        columnas = ["Preventista", "Cant. Facturas", "Total Vendido"]
        return self._exportar_excel(datos, columnas, ruta_salida, "Vtas. Preventista")